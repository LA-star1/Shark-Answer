"""Quick tests for each pipeline — run independently.

Usage:
    # Test all (no API calls — unit tests only)
    pytest test_quick.py -v

    # Test with live API calls (requires .env with keys)
    pytest test_quick.py -v -m live

    # Test a specific pipeline
    pytest test_quick.py -v -k "pipeline_a"
"""

from __future__ import annotations

import asyncio
import json
import os
import pytest

# ===== Unit Tests (no API calls) =====


class TestConfig:
    def test_subject_pipeline_mapping(self):
        from shark_answer.config import Subject, Pipeline, SUBJECT_PIPELINE_MAP
        assert SUBJECT_PIPELINE_MAP[Subject.PHYSICS] == Pipeline.SCIENCE_MATH
        assert SUBJECT_PIPELINE_MAP[Subject.ECONOMICS] == Pipeline.ESSAY
        assert SUBJECT_PIPELINE_MAP[Subject.COMPUTER_SCIENCE] == Pipeline.CS

    def test_app_config_from_env(self):
        from shark_answer.config import AppConfig
        config = AppConfig.from_env()
        assert config.max_answer_versions == 5
        assert config.log_level in ("INFO", "DEBUG", "WARNING", "ERROR")

    def test_model_cost_defined(self):
        from shark_answer.config import MODEL_COST_PER_M_TOKENS, ModelProvider
        for provider in ModelProvider:
            assert provider in MODEL_COST_PER_M_TOKENS, f"Missing cost for {provider.value}"

    def test_pipeline_config_structure(self):
        from shark_answer.config import PIPELINE_CONFIG, ModelProvider
        # Pipeline A
        assert "A_science" in PIPELINE_CONFIG
        assert ModelProvider.O3PRO in PIPELINE_CONFIG["A_science"]["solvers"]
        assert ModelProvider.GLM in PIPELINE_CONFIG["A_science"]["solvers"]
        assert PIPELINE_CONFIG["A_science"]["judge"] == ModelProvider.CLAUDE
        # Pipeline B
        assert "B_essay" in PIPELINE_CONFIG
        angles = PIPELINE_CONFIG["B_essay"]["angles"]
        assert ModelProvider.CLAUDE in angles
        assert ModelProvider.GPT4O in angles
        assert ModelProvider.GLM in angles
        assert ModelProvider.KIMI in angles
        # all 7 angle models present
        assert len(angles) == 7
        # Pipeline C
        assert "C_cs" in PIPELINE_CONFIG
        assert ModelProvider.MINIMAX in PIPELINE_CONFIG["C_cs"]["solvers"]
        assert ModelProvider.GLM in PIPELINE_CONFIG["C_cs"]["solvers"]

    def test_o3pro_uses_openai_key(self):
        """O3PRO and GPT4O share the same OPENAI_API_KEY env var."""
        from shark_answer.config import AppConfig, ModelProvider
        import os
        os.environ.setdefault("OPENAI_API_KEY", "sk-test-key")
        config = AppConfig.from_env()
        if ModelProvider.O3PRO in config.models:
            o3 = config.models[ModelProvider.O3PRO]
            gpt = config.models.get(ModelProvider.GPT4O)
            if gpt:
                assert o3.api_key == gpt.api_key, "O3PRO and GPT4O must share OPENAI_API_KEY"

    def test_glm_and_o3pro_in_model_provider_enum(self):
        from shark_answer.config import ModelProvider
        assert ModelProvider.GLM == "glm"
        assert ModelProvider.O3PRO == "o3pro"

    def test_registry_knows_glm_and_o3pro_models(self):
        from shark_answer.providers.registry import DEFAULT_MODELS
        from shark_answer.config import ModelProvider
        assert ModelProvider.GLM in DEFAULT_MODELS
        assert ModelProvider.O3PRO in DEFAULT_MODELS
        assert DEFAULT_MODELS[ModelProvider.O3PRO] == "o3-pro"
        assert DEFAULT_MODELS[ModelProvider.GLM] == "glm-5"


class TestCostTracker:
    def test_record_and_summary(self):
        from shark_answer.utils.cost_tracker import CostTracker
        from shark_answer.providers.base import ModelResponse, TokenUsage

        tracker = CostTracker(budget_warning_usd=100.0)
        resp = ModelResponse(
            content="test",
            provider="claude",
            model_name="claude-sonnet-4-20250514",
            usage=TokenUsage(input_tokens=1000, output_tokens=500),
            success=True,
        )
        tracker.record(resp, subject="physics", pipeline="A")

        summary = tracker.summary()
        assert summary["total_calls"] == 1
        assert summary["total_input_tokens"] == 1000
        assert summary["total_output_tokens"] == 500
        assert summary["total_cost_usd"] > 0
        assert "claude" in summary["by_provider"]
        assert "physics" in summary["by_subject"]


class TestMathVerifier:
    def test_extract_numeric_answer(self):
        from shark_answer.utils.math_verifier import extract_numeric_answer
        assert extract_numeric_answer("The answer is 42.5 m/s") == 42.5
        assert extract_numeric_answer("= 3.14") == 3.14
        assert extract_numeric_answer("\\boxed{100}") == 100
        assert extract_numeric_answer("no number here") is None

    def test_verify_numeric_agreement(self):
        from shark_answer.utils.math_verifier import verify_numeric_agreement
        agreed, vals = verify_numeric_agreement([
            "The answer = 42.5",
            "Result is 42.5",
            "= 42.5 m/s",
        ])
        assert agreed is True
        assert len(vals) == 3

    def test_verify_numeric_disagreement(self):
        from shark_answer.utils.math_verifier import verify_numeric_agreement
        agreed, vals = verify_numeric_agreement([
            "answer = 42.5",
            "answer = 50.0",
        ])
        assert agreed is False

    def test_verify_physics_calculation(self):
        from shark_answer.utils.math_verifier import verify_physics_calculation
        result = verify_physics_calculation(
            formula="0.5 * m * v**2",
            variables={"m": 2.0, "v": 3.0},
            expected_result=9.0,
        )
        assert result.verified is True

    def test_verify_sympy(self):
        from shark_answer.utils.math_verifier import verify_with_sympy
        result = verify_with_sympy("2 + 3", "5")
        assert result.verified is True


class TestImageExtractor:
    def test_parse_extraction_response(self):
        from shark_answer.utils.image_extractor import _parse_extraction_response
        response = """```json
[
  {
    "number": "1(a)",
    "text": "Calculate the velocity when $v = u + at$",
    "marks": 3,
    "has_diagram": false,
    "diagram_description": "",
    "question_type": "calculation",
    "topic_hints": ["kinematics"]
  }
]
```"""
        questions = _parse_extraction_response(response)
        assert len(questions) == 1
        assert questions[0].number == "1(a)"
        assert questions[0].marks == 3
        assert questions[0].question_type == "calculation"


class TestExaminerProfile:
    def test_default_profiles_load(self, tmp_path):
        from shark_answer.modules.examiner_profile import ExaminerProfileManager
        manager = ExaminerProfileManager(tmp_path / "profiles")
        profiles = manager.list_profiles()
        assert len(profiles) >= 1  # at least default profiles

    def test_profile_prompt_guidance(self):
        from shark_answer.modules.examiner_profile import ExaminerProfile
        profile = ExaminerProfile(
            name="Test",
            subject="physics",
            evaluation_depth=0.9,
            strict_on_units=True,
        )
        guidance = profile.to_prompt_guidance()
        assert "DEEP evaluation" in guidance
        assert "units" in guidance.lower()

    def test_crud_profile(self, tmp_path):
        from shark_answer.modules.examiner_profile import (
            ExaminerProfileManager, ExaminerProfile,
        )
        manager = ExaminerProfileManager(tmp_path / "profiles")
        p = ExaminerProfile(name="Custom", subject="economics", region="UK")
        manager.add_profile(p)
        assert manager.get_profile("Custom") is not None
        manager.delete_profile("Custom")
        assert manager.get_profile("Custom") is None


class TestKnowledgeBase:
    def test_add_and_retrieve(self, tmp_path):
        from shark_answer.knowledge_base.store import KnowledgeBase, MarkSchemeCriteria
        kb = KnowledgeBase(tmp_path / "kb")
        ms = MarkSchemeCriteria(
            subject="physics", paper="9702/42", year="2024",
            session="May/June", topic="Kinematics",
            question_number="1(a)(i)", marks=3,
            marking_points=["Uses v=u+at", "Substitutes correctly", "Correct answer with units"],
        )
        kb.add_mark_scheme(ms)
        results = kb.get_mark_scheme("physics", topic="kinematics")
        assert len(results) == 1
        assert results[0].marking_points[0] == "Uses v=u+at"

    def test_marking_context_string(self, tmp_path):
        from shark_answer.knowledge_base.store import KnowledgeBase, MarkSchemeCriteria
        kb = KnowledgeBase(tmp_path / "kb")
        ms = MarkSchemeCriteria(
            subject="physics", paper="9702/42", year="2024",
            session="May/June", topic="Kinematics",
            question_number="1(a)", marks=5,
            marking_points=["Point A", "Point B"],
            common_errors=["Unit error"],
        )
        kb.add_mark_scheme(ms)
        ctx = kb.get_marking_context("physics", "kinematics")
        assert "Point A" in ctx
        assert "Unit error" in ctx


class TestPipelineRouter:
    def test_physics_routes_to_pipeline_a(self):
        from shark_answer.pipelines.router import route_question
        from shark_answer.config import Subject, Pipeline
        from shark_answer.utils.image_extractor import ExtractedQuestion

        q = ExtractedQuestion(number="1", text="Calculate force", question_type="calculation")
        assert route_question(q, Subject.PHYSICS) == Pipeline.SCIENCE_MATH

    def test_economics_routes_to_pipeline_b(self):
        from shark_answer.pipelines.router import route_question
        from shark_answer.config import Subject, Pipeline
        from shark_answer.utils.image_extractor import ExtractedQuestion

        q = ExtractedQuestion(number="1", text="Discuss market failure", question_type="essay")
        assert route_question(q, Subject.ECONOMICS) == Pipeline.ESSAY

    def test_cs_routes_to_pipeline_c(self):
        from shark_answer.pipelines.router import route_question
        from shark_answer.config import Subject, Pipeline
        from shark_answer.utils.image_extractor import ExtractedQuestion

        q = ExtractedQuestion(number="1", text="Write pseudocode", question_type="code")
        assert route_question(q, Subject.COMPUTER_SCIENCE) == Pipeline.CS

    def test_bio_essay_routes_to_pipeline_b(self):
        from shark_answer.pipelines.router import route_question
        from shark_answer.config import Subject, Pipeline
        from shark_answer.utils.image_extractor import ExtractedQuestion

        q = ExtractedQuestion(number="1", text="Discuss enzyme action", question_type="essay")
        assert route_question(q, Subject.BIOLOGY) == Pipeline.ESSAY

    def test_bio_calculation_routes_to_pipeline_a(self):
        from shark_answer.pipelines.router import route_question
        from shark_answer.config import Subject, Pipeline
        from shark_answer.utils.image_extractor import ExtractedQuestion

        q = ExtractedQuestion(number="1", text="Calculate magnification", question_type="calculation")
        assert route_question(q, Subject.BIOLOGY) == Pipeline.SCIENCE_MATH


class TestExplanationPrompts:
    def test_science_prompt_includes_steps(self):
        from shark_answer.modules.explanation import build_explanation_prompt
        from shark_answer.config import Pipeline
        prompt = build_explanation_prompt(
            Pipeline.SCIENCE_MATH, "Calculate F=ma", "F = 10*2 = 20N"
        )
        assert "Step-by-Step" in prompt
        assert "Common Mistakes" in prompt

    def test_essay_prompt_includes_criteria(self):
        from shark_answer.modules.explanation import build_explanation_prompt
        from shark_answer.config import Pipeline
        prompt = build_explanation_prompt(
            Pipeline.ESSAY, "Discuss inflation", "Essay text here"
        )
        assert "Mark Scheme" in prompt
        assert "Argument Structure" in prompt

    def test_cs_prompt_includes_complexity(self):
        from shark_answer.modules.explanation import build_explanation_prompt
        from shark_answer.config import Pipeline
        prompt = build_explanation_prompt(
            Pipeline.CS, "Write sort algorithm", "PROCEDURE BubbleSort"
        )
        assert "Complexity" in prompt

    def test_chinese_language(self):
        from shark_answer.modules.explanation import build_explanation_prompt
        from shark_answer.config import Pipeline
        prompt = build_explanation_prompt(
            Pipeline.SCIENCE_MATH, "Q", "A", language="zh"
        )
        assert "Chinese" in prompt or "中文" in prompt


class TestPDFExport:
    def test_generate_pdf(self):
        from shark_answer.app import _generate_pdf
        data = {
            "total_questions": 1,
            "cost_summary": {"total_cost_usd": 0.01},
            "results": [{
                "question_number": "1(a)",
                "question_text": "Calculate F=ma when m=10kg, a=2m/s^2",
                "pipeline": "A",
                "subject": "physics",
                "versions": [{
                    "version_number": 1,
                    "answer_text": "F = ma = 10 * 2 = 20 N",
                    "explanation_text": "Step 1: Identify formula",
                    "approach_label": "Direct calculation",
                    "provider": "claude",
                    "verified": True,
                    "quality_score": None,
                }],
                "verification_notes": "",
                "disagreement_resolved": True,
                "errors": [],
            }],
        }
        pdf_bytes = _generate_pdf(data)
        assert len(pdf_bytes) > 100
        assert pdf_bytes[:5] == b'%PDF-'

    def test_generate_pdf_with_special_chars(self):
        from shark_answer.app import _generate_pdf
        data = {
            "total_questions": 1,
            "cost_summary": {"total_cost_usd": 0.0},
            "results": [{
                "question_number": "2",
                "question_text": "What is <x> & y > z?",
                "pipeline": "A",
                "subject": "math",
                "versions": [{
                    "version_number": 1,
                    "answer_text": "x < y & z > 0\nLine 2 with $\\alpha$",
                    "explanation_text": "",
                    "approach_label": "Test",
                    "provider": "gpt4o",
                    "verified": False,
                    "quality_score": 85.0,
                }],
                "verification_notes": "",
                "disagreement_resolved": True,
                "errors": [],
            }],
        }
        pdf_bytes = _generate_pdf(data)
        assert pdf_bytes[:5] == b'%PDF-'


class TestDOCXExport:
    def test_generate_docx(self):
        from shark_answer.app import _generate_docx
        data = {
            "total_questions": 1,
            "cost_summary": {"total_cost_usd": 0.02},
            "results": [{
                "question_number": "3",
                "question_text": "Discuss market failure",
                "pipeline": "B",
                "subject": "economics",
                "versions": [{
                    "version_number": 1,
                    "answer_text": "Market failure occurs when...",
                    "explanation_text": "The argument structure is...",
                    "approach_label": "Orthodox approach",
                    "provider": "claude",
                    "verified": False,
                    "quality_score": 78.0,
                }],
                "verification_notes": "",
                "disagreement_resolved": True,
                "errors": [],
            }],
        }
        docx_bytes = _generate_docx(data)
        assert len(docx_bytes) > 100
        # DOCX is a ZIP file, starts with PK signature
        assert docx_bytes[:2] == b'PK'


class TestFrontendEndpoints:
    """Test that HTML pages and new API endpoints work."""

    @pytest.mark.asyncio
    async def test_index_page_serves_html(self):
        from httpx import AsyncClient, ASGITransport
        from shark_answer.app import app

        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as client:
            resp = await client.get("/")
            assert resp.status_code == 200
            assert "Shark Answer" in resp.text
            assert "dropZone" in resp.text

    @pytest.mark.asyncio
    async def test_history_endpoint_empty(self):
        from httpx import AsyncClient, ASGITransport
        from shark_answer.app import app

        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as client:
            resp = await client.get("/api/history")
            assert resp.status_code == 200
            assert isinstance(resp.json(), list)

    @pytest.mark.asyncio
    async def test_export_pdf_endpoint(self):
        from httpx import AsyncClient, ASGITransport
        from shark_answer.app import app

        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as client:
            resp = await client.post("/api/export/pdf", json={
                "data": {
                    "total_questions": 1,
                    "cost_summary": {"total_cost_usd": 0.0},
                    "results": [{
                        "question_number": "1",
                        "question_text": "Test Q",
                        "pipeline": "A",
                        "subject": "physics",
                        "versions": [{
                            "version_number": 1,
                            "answer_text": "Test A",
                            "explanation_text": "",
                            "approach_label": "Test",
                            "provider": "claude",
                            "verified": False,
                            "quality_score": None,
                        }],
                        "verification_notes": "",
                        "disagreement_resolved": True,
                        "errors": [],
                    }],
                },
            })
            assert resp.status_code == 200
            assert resp.headers["content-type"] == "application/pdf"

    @pytest.mark.asyncio
    async def test_export_docx_endpoint(self):
        from httpx import AsyncClient, ASGITransport
        from shark_answer.app import app

        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as client:
            resp = await client.post("/api/export/docx", json={
                "data": {
                    "total_questions": 1,
                    "cost_summary": {"total_cost_usd": 0.0},
                    "results": [{
                        "question_number": "1",
                        "question_text": "Test Q",
                        "pipeline": "A",
                        "subject": "physics",
                        "versions": [{
                            "version_number": 1,
                            "answer_text": "Test A",
                            "explanation_text": "",
                            "approach_label": "Test",
                            "provider": "claude",
                            "verified": False,
                            "quality_score": None,
                        }],
                        "verification_notes": "",
                        "disagreement_resolved": True,
                        "errors": [],
                    }],
                },
            })
            assert resp.status_code == 200
            assert "wordprocessingml" in resp.headers["content-type"]


# ===== Live API Tests (requires API keys) =====

def _has_api_keys() -> bool:
    return bool(os.getenv("ANTHROPIC_API_KEY") or os.getenv("OPENAI_API_KEY"))


live = pytest.mark.skipif(
    not _has_api_keys(),
    reason="No API keys configured — set ANTHROPIC_API_KEY or OPENAI_API_KEY",
)


@live
class TestLivePipelineA:
    """Live test for Pipeline A with real API calls."""

    @pytest.mark.asyncio
    async def test_simple_physics_question(self):
        from shark_answer.config import AppConfig, Pipeline
        from shark_answer.providers.registry import ProviderRegistry
        from shark_answer.utils.cost_tracker import CostTracker
        from shark_answer.utils.image_extractor import ExtractedQuestion
        from shark_answer.pipelines.pipeline_a_science import run_pipeline_a

        config = AppConfig.from_env()
        registry = ProviderRegistry(config)
        tracker = CostTracker()

        q = ExtractedQuestion(
            number="1(a)",
            text="A car accelerates uniformly from rest to 20 m/s in 5 seconds. Calculate the acceleration.",
            marks=2,
            question_type="calculation",
            topic_hints=["kinematics"],
        )

        result = await run_pipeline_a(
            question=q, subject="physics", registry=registry,
            config=config, cost_tracker=tracker,
            max_versions=2,
        )

        assert len(result.versions) >= 1, f"Errors: {result.errors}"
        assert "4" in result.versions[0].answer_text  # a = 20/5 = 4 m/s²
        print(f"\nCost: ${tracker.total_cost:.4f}")
        print(f"Answer preview: {result.versions[0].answer_text[:200]}")


@live
class TestLivePipelineB:
    """Live test for Pipeline B with real API calls."""

    @pytest.mark.asyncio
    async def test_simple_economics_essay(self):
        from shark_answer.config import AppConfig, Pipeline
        from shark_answer.providers.registry import ProviderRegistry
        from shark_answer.utils.cost_tracker import CostTracker
        from shark_answer.utils.image_extractor import ExtractedQuestion
        from shark_answer.pipelines.pipeline_b_essay import run_pipeline_b

        config = AppConfig.from_env()
        registry = ProviderRegistry(config)
        tracker = CostTracker()

        q = ExtractedQuestion(
            number="4(a)",
            text="Discuss whether a government should intervene to correct market failure caused by negative externalities.",
            marks=12,
            question_type="essay",
            topic_hints=["market failure", "externalities"],
        )

        result = await run_pipeline_b(
            question=q, subject="economics", registry=registry,
            config=config, cost_tracker=tracker,
            max_versions=2,  # keep costs low for testing
        )

        assert len(result.versions) >= 1, f"Errors: {result.errors}"
        print(f"\nCost: ${tracker.total_cost:.4f}")
        print(f"Versions generated: {len(result.versions)}")
        for v in result.versions:
            print(f"  V{v.version_number}: {v.approach_label} ({v.provider}) "
                  f"score={v.quality_score}")


@live
class TestLivePipelineC:
    """Live test for Pipeline C with real API calls."""

    @pytest.mark.asyncio
    async def test_simple_cs_question(self):
        from shark_answer.config import AppConfig, Pipeline
        from shark_answer.providers.registry import ProviderRegistry
        from shark_answer.utils.cost_tracker import CostTracker
        from shark_answer.utils.image_extractor import ExtractedQuestion
        from shark_answer.pipelines.pipeline_c_cs import run_pipeline_c

        config = AppConfig.from_env()
        registry = ProviderRegistry(config)
        tracker = CostTracker()

        q = ExtractedQuestion(
            number="3(b)",
            text="Write a pseudocode procedure that takes an array of integers and returns the largest value. The procedure should use a loop to iterate through the array.",
            marks=5,
            question_type="code",
            topic_hints=["arrays", "iteration"],
        )

        result = await run_pipeline_c(
            question=q, subject="computer_science", registry=registry,
            config=config, cost_tracker=tracker,
            max_versions=2,
        )

        assert len(result.versions) >= 1, f"Errors: {result.errors}"
        print(f"\nCost: ${tracker.total_cost:.4f}")
        for v in result.versions:
            print(f"  V{v.version_number}: verified={v.verified} ({v.provider})")


class TestFileConverter:
    """Unit tests for the file format converter utility."""

    def test_native_image_passthrough(self):
        from shark_answer.utils.file_converter import convert_file_to_images
        fake_png = b'\x89PNG\r\n\x1a\n' + b'\x00' * 100
        result = convert_file_to_images("exam.png", fake_png)
        assert result == [fake_png]

    def test_jpeg_passthrough(self):
        from shark_answer.utils.file_converter import convert_file_to_images
        fake_jpg = b'\xff\xd8\xff' + b'\x00' * 100
        result = convert_file_to_images("scan.jpg", fake_jpg)
        assert result == [fake_jpg]

    def test_pdf_conversion_returns_list(self):
        """Test that a minimal valid PDF is accepted (empty list or images)."""
        from shark_answer.utils.file_converter import convert_file_to_images
        # Minimal PDF-like bytes — will fail gracefully
        fake_pdf = b'%PDF-1.4\n%%EOF'
        result = convert_file_to_images("paper.pdf", fake_pdf)
        # Should return a list (empty on parse error is acceptable)
        assert isinstance(result, list)

    def test_unknown_extension_passthrough(self):
        from shark_answer.utils.file_converter import convert_file_to_images
        data = b'\x00\x01\x02\x03'
        result = convert_file_to_images("mystery.xyz", data)
        assert result == [data]

    def test_text_to_png_renders(self):
        from shark_answer.utils.file_converter import _text_to_png
        png = _text_to_png("Hello, World!\nLine 2")
        if png is not None:  # Pillow might not be installed in CI
            assert png[:4] == b'\x89PNG'

    def test_docx_conversion_empty(self):
        """Empty DOCX-like bytes should return empty list (graceful failure)."""
        from shark_answer.utils.file_converter import convert_file_to_images
        result = convert_file_to_images("empty.docx", b'not a real docx')
        assert isinstance(result, list)


class TestMarkdownExport:
    """Tests for Markdown export generation."""

    def _sample_data(self):
        return {
            "total_questions": 1,
            "cost_summary": {"total_cost_usd": 0.005},
            "results": [{
                "question_number": "1(a)",
                "question_text": "Describe Newton's second law",
                "pipeline": "A",
                "subject": "physics",
                "versions": [{
                    "version_number": 1,
                    "answer_text": "F = ma where F is force, m is mass, a is acceleration",
                    "explanation_text": "This follows from first principles",
                    "approach_label": "Direct definition",
                    "provider": "claude",
                    "verified": True,
                    "quality_score": 95.0,
                    "language": "en",
                }],
                "verification_notes": "",
                "disagreement_resolved": True,
                "errors": [],
            }],
        }

    def test_markdown_contains_question(self):
        from shark_answer.app import _generate_markdown
        md = _generate_markdown(self._sample_data())
        assert "Question 1(a)" in md
        assert "Newton" in md

    def test_markdown_contains_answer(self):
        from shark_answer.app import _generate_markdown
        md = _generate_markdown(self._sample_data())
        assert "F = ma" in md

    def test_markdown_has_headers(self):
        from shark_answer.app import _generate_markdown
        md = _generate_markdown(self._sample_data())
        assert md.startswith("# Shark Answer")

    def test_markdown_verified_badge(self):
        from shark_answer.app import _generate_markdown
        md = _generate_markdown(self._sample_data())
        assert "✓" in md

    def test_txt_export(self):
        from shark_answer.app import _generate_txt
        txt = _generate_txt(self._sample_data())
        assert "SHARK ANSWER" in txt
        assert "Question 1(a)" in txt.upper() or "QUESTION 1(A)" in txt.upper()
        assert "F = ma" in txt

    def test_xlsx_export(self):
        from shark_answer.app import _generate_xlsx
        xlsx_bytes = _generate_xlsx(self._sample_data())
        assert len(xlsx_bytes) > 100
        # XLSX is a ZIP, starts with PK
        assert xlsx_bytes[:2] == b'PK'

    def test_xlsx_has_content(self):
        """Load the generated XLSX and verify row content."""
        from shark_answer.app import _generate_xlsx
        import io
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not installed")
        xlsx_bytes = _generate_xlsx(self._sample_data())
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        # Row 1 = headers, row 2 = data
        assert ws.max_row >= 2
        # Check question number in column A row 2
        cell_val = ws.cell(row=2, column=1).value
        assert cell_val == "1(a)"


class TestNewExportEndpoints:
    """Test new export endpoints via FastAPI test client."""

    def _sample_payload(self):
        return {
            "data": {
                "total_questions": 1,
                "cost_summary": {"total_cost_usd": 0.001},
                "results": [{
                    "question_number": "2",
                    "question_text": "Explain photosynthesis",
                    "pipeline": "B",
                    "subject": "biology",
                    "versions": [{
                        "version_number": 1,
                        "answer_text": "Photosynthesis converts CO2 and water into glucose using light energy.",
                        "explanation_text": "This is a fundamental biological process.",
                        "approach_label": "Standard explanation",
                        "provider": "claude",
                        "verified": False,
                        "quality_score": 82.0,
                        "language": "en",
                    }],
                    "verification_notes": "",
                    "disagreement_resolved": True,
                    "errors": [],
                }],
            }
        }

    @pytest.mark.asyncio
    async def test_export_markdown_endpoint(self):
        from httpx import AsyncClient, ASGITransport
        from shark_answer.app import app

        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as client:
            resp = await client.post("/api/export/md", json=self._sample_payload())
            assert resp.status_code == 200
            assert "markdown" in resp.headers["content-type"]
            assert "Shark Answer" in resp.text

    @pytest.mark.asyncio
    async def test_export_txt_endpoint(self):
        from httpx import AsyncClient, ASGITransport
        from shark_answer.app import app

        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as client:
            resp = await client.post("/api/export/txt", json=self._sample_payload())
            assert resp.status_code == 200
            assert "text/plain" in resp.headers["content-type"]
            assert "SHARK ANSWER" in resp.text

    @pytest.mark.asyncio
    async def test_export_xlsx_endpoint(self):
        from httpx import AsyncClient, ASGITransport
        from shark_answer.app import app

        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as client:
            resp = await client.post("/api/export/xlsx", json=self._sample_payload())
            assert resp.status_code == 200
            assert "spreadsheetml" in resp.headers["content-type"]
            assert resp.content[:2] == b'PK'


class TestChatEndpoint:
    """Tests for the answer correction chat endpoint."""

    @pytest.mark.asyncio
    async def test_chat_missing_submission(self):
        from httpx import AsyncClient, ASGITransport
        from shark_answer.app import app

        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as client:
            resp = await client.post("/api/chat", json={
                "submission_id": "nonexistent",
                "question_number": "1",
                "message": "Is this correct?",
            })
            assert resp.status_code == 404

    @pytest.mark.asyncio
    async def test_chat_history_endpoint_empty(self):
        from httpx import AsyncClient, ASGITransport
        from shark_answer.app import app

        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as client:
            resp = await client.get("/api/chat/unknown_id/1")
            # Should return empty list or 200 with []
            assert resp.status_code == 200
            assert resp.json() == []

    @pytest.mark.asyncio
    async def test_index_contains_chat_panel(self):
        from httpx import AsyncClient, ASGITransport
        from shark_answer.app import app

        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as client:
            resp = await client.get("/")
            assert resp.status_code == 200
            assert "chatSidebar" in resp.text
            assert "Ask Claude" in resp.text

    @pytest.mark.asyncio
    async def test_index_accepts_pdf_docx(self):
        """Verify the new file input accepts PDF and DOCX."""
        from httpx import AsyncClient, ASGITransport
        from shark_answer.app import app

        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as client:
            resp = await client.get("/")
            assert resp.status_code == 200
            assert ".pdf" in resp.text
            assert ".docx" in resp.text

    @pytest.mark.asyncio
    async def test_index_has_new_export_buttons(self):
        """Verify MD, TXT, XLSX export buttons appear in HTML."""
        from httpx import AsyncClient, ASGITransport
        from shark_answer.app import app

        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as client:
            resp = await client.get("/")
            assert resp.status_code == 200
            assert "exportMarkdown" in resp.text
            assert "exportTxt" in resp.text
            assert "exportXlsx" in resp.text


@live
class TestLiveAPI:
    """Live test for the FastAPI endpoints."""

    @pytest.mark.asyncio
    async def test_health_endpoint(self):
        from httpx import AsyncClient, ASGITransport
        from shark_answer.app import app

        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as client:
            resp = await client.get("/api/health")
            assert resp.status_code == 200
            data = resp.json()
            assert data["status"] == "ok"

    @pytest.mark.asyncio
    async def test_debug_models_endpoint_exists(self):
        """Verify /api/debug/models endpoint returns expected structure."""
        from httpx import AsyncClient, ASGITransport
        from shark_answer.app import app

        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as client:
            resp = await client.get("/api/debug/models")
            assert resp.status_code == 200
            data = resp.json()
            assert "models" in data
            assert "total_configured" in data
            assert "total_ok" in data

    @pytest.mark.asyncio
    async def test_examiner_profiles_endpoint(self):
        from httpx import AsyncClient, ASGITransport
        from shark_answer.app import app

        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as client:
            resp = await client.get("/api/examiner/profiles")
            assert resp.status_code == 200
            profiles = resp.json()
            assert isinstance(profiles, list)
